from pathlib import Path

import pytest

from src.document_loader import DocumentLoader


def test_load_docx(tmp_path: Path) -> None:
    from docx import Document as DocxDocument

    d = DocxDocument()
    d.add_paragraph("UniqueDocxMarkerAlpha")
    p = tmp_path / "t.docx"
    d.save(p)
    loader = DocumentLoader(chunk_size=80, chunk_overlap=5)
    docs = loader.load_documents(str(p))
    joined = " ".join(x.page_content for x in docs)
    assert "UniqueDocxMarkerAlpha" in joined


def test_load_xlsx(tmp_path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "UniqueXlsxMarkerBeta"
    ws["B1"] = 42
    p = tmp_path / "t.xlsx"
    wb.save(p)
    loader = DocumentLoader(chunk_size=100, chunk_overlap=10)
    docs = loader.load_documents(str(p))
    joined = " ".join(x.page_content for x in docs)
    assert "UniqueXlsxMarkerBeta" in joined


def test_legacy_doc_suffix_rejected(tmp_path: Path) -> None:
    p = tmp_path / "legacy.doc"
    p.write_bytes(b"dummy")
    loader = DocumentLoader()
    with pytest.raises(ValueError, match="Legacy Word"):
        loader.load_documents(str(p))
